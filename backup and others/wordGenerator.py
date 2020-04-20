from xml.etree import ElementTree as ET
import openpyxl
import os.path
import re
# from googletrans import Translator
from openpyxl.styles import Font
from openpyxl.styles.colors import RED
# https://www.v2ex.com/t/290981
from translation import baidu, google, youdao, bing


def translate(str, name):
    if name == "google":
        return google(str, dst='zh-CN', proxies={'http': '127.0.0.1:1080'})
    elif name == "baidu":
        return baidu(str, dst='zh-CN')
    elif name == "youdao":
        return youdao(str, dst='zh-CN')
    # elif name == "icibar":
    #     return icibar(str, dst='zh')


def write_xml(file_name, ret, trans_name):
    file_name = file_name+".xlsx"
    if os.path.isfile(file_name):
        os.remove(file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    i = 1
    # trans_list = []

    trans = ""
    trans_result = ""
    # translator = Translator(to_lang="chinese")

    for (key, value) in ret.items():
        # ws.cell(i, 1, key)
        # ws.cell(i, 2, value)
        trans = trans+key+". "
        if(i % 100 == 0):  # google translator一次只允许翻译5000个字符，一般单词少于20个字符
            trans_result = trans_result + \
                translate(trans, trans_name)
            trans = ""
        i = i+1

    if trans != "":
        trans_result = trans_result + \
            translate(trans, trans_name)

    # 翻译后变成中文逗号，不用空格分隔，避免上下文联系，，，，实际上更复杂，试验后发现使用句号上下文联系最少，如果使用特殊符号，google翻译会自动清楚，怪google太智能 orz

    # trans_result = re.sub('[\n\t.]+', '', trans_result)
    trans_result_list = re.split('。|！|？|·', trans_result)
    # trans_result_list.append("x")
    # trans_result_list.append("x")
    print(len(trans_result_list))
    print(trans_result_list)
    print(len(ret))

    tree = ET.parse("data/udicWord.xml")
    udic = set()
    for elem in tree.iter(tag='CustomizeListItem'):
        udic.add(elem.get("word"))

    color_font = Font(color=RED)

    i = 0
    index = 1
    for (key, value) in ret.items():
        # if key[-1] == 's':  # 删除复数形式，但是，也有不足，可能复数形式出现次数更多，排在前面
        #     key_temp = key[:-1]
        #     if key_temp in ret:
        #         i += 1
        #         continue
        temp = trans_result_list[i]
        if not bool(re.search('[a-zA-Z\-]+', temp)):  # 翻译中含有英文，一般是没有翻译出，或者其他问题
            ws.cell(index, 1, key)
            ws.cell(index, 2, value)
            ws.cell(index, 3, temp)
            if key in udic:
                ws.cell(index, 1).font = color_font
                ws.cell(index, 2).font = color_font
                ws.cell(index, 3).font = color_font
            index += 1
        i = i+1
        if i >= len(trans_result_list):
            break

    print("一共有"+str(index)+"个单词")
    wb.save(filename=(file_name))


def auto_generate_word(source_file, delete_file,  filter_times):
    text = dict()
    file = open("data/"+source_file, "r")
    file_text = file.read()
    file_text = file_text.lower()

    file_text = re.sub('-[\n\t]+', '', file_text)
    word_list = re.findall('[a-zA-Z\-\'][a-zA-Z\-\'][a-zA-Z\-\'][a-zA-Z\-\']+',
                           file_text)  # 至少4个字符的才能入选，一个单词和两个单词太简单，而且容易导致翻译依赖，比如：etc.
    for word in word_list:
        if word[0] == "\'" and word[-1] == "\'":  # 防止出现 'hello'的形式
            word = word[1:-1]
        if word not in text:
            text[word] = 1
        else:
            text[word] += 1
    # print(text)

    file2 = open("data/"+delete_file, "r")
    delete_text = file2.read()
    delete_word_list = re.findall('[a-zA-Z\-\']+', delete_text)
    delete_text = set()
    for delete_word in delete_word_list:
        delete_text.add(delete_word)

    file3 = open("data/firstname_all.txt", "r")
    delete_name_text = file3.read()
    file4 = open("data/surname_all.txt", "r")
    delete_name_text = delete_name_text+"\n"+file4.read()
    # print(delete_name_text)

    delete_name_text = delete_name_text.lower()
    # delete_name_list = re.findall('[a-zA-Z\-\']+', delete_name_text)
    delete_name_list = delete_name_text.split("\n")
    delete_name = set()
    for temp in delete_name_list:
        delete_name.add(temp)

    tree_city = ET.parse("data/en_city.xml")
    for elem in tree_city.iter(tag='CountryRegion'):
        delete_name.add(elem.get("Name").lower())
    for elem in tree_city.iter(tag='City'):
        delete_name.add(elem.get("Name").lower())

    file5 = open("data/simple word lists.txt", "r")
    delete_simple_words_text = file5.read()
    delete_simple_words_list = delete_simple_words_text .split("\n")
    delete_simple_words = set()
    for temp in delete_simple_words_list:
        delete_simple_words.add(temp)
    # print(delete_simple_words)

    # print(delete_name)

    ret = dict()
    for (key, value) in text.items():
        if key not in delete_text and key not in delete_name and key not in delete_simple_words and value > filter_times:
            ret[key] = value

    file.close()
    file2.close()
    file3.close()
    file4.close()
    file5.close()
    return ret


def main():
    # need set
    ##############################
    # is_use_eudic = False
    source_file = "17.txt"
    # source_file = "thinking in java.txt"
    delete_file = "delete.txt"
    filter_times = 0  # 出现频率小于等于3不出现
    trans_name = "youdao"
    ##############################
    # print(google('hello world!', dst='zh-CN'))
    # print(google('hello world!', dst='ru'))
    # print(baidu('hello world!', dst='zh'))
    # print(baidu('hello world!', dst='ru'))
    print(bing('hello world!', dst='zh-CHS'))

    print("============================================")
    print("源文件："+source_file)
    print("不显示的常用单词表："+delete_file)
    print("删除简单单词：data/simple word lists.txt")
    print("高亮欧路词典标记单词")
    print("只显示出现"+str(filter_times)+"以上的单词")
    print("使用"+trans_name+"翻译")
    # print("是否只匹配给定词典："+str(is_use_eudic))
    print("============================================")
    print("生成单词列表中...")
    ret = auto_generate_word(source_file, delete_file,
                             filter_times)
    print("翻译和写入Excel文件...")
    write_xml("word_lists", ret, trans_name)
    print("完成！")


if __name__ == "__main__":
    main()
