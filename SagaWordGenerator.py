# -*- coding:utf-8 -*-
import sys
import re
import collections
import nltk
from nltk.stem.wordnet import WordNetLemmatizer
from nltk.tokenize import word_tokenize
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import sys
from urllib.parse import urlparse, quote, urlencode, unquote
from urllib.request import urlopen
import re
import os.path
from openpyxl.styles import Font
from openpyxl.styles.colors import RED

# patterns that used to find or/and replace particular chars or words
# to find chars that are not a letter, a blank or a quotation
pat_letter = re.compile(r'[^a-zA-Z \']+')
# to find the 's following the pronouns. re.I is refers to ignore case
pat_is = re.compile("(it|he|she|that|this|there|here)(\'s)", re.I)
# to find the 's following the letters
pat_s = re.compile("(?<=[a-zA-Z])\'s")
# to find the ' following the words ending by s
pat_s2 = re.compile("(?<=s)\'s?")
# to find the abbreviation of not
pat_not = re.compile("(?<=[a-zA-Z])n\'t")
# to find the abbreviation of would
pat_would = re.compile("(?<=[a-zA-Z])\'d")
# to find the abbreviation of will
pat_will = re.compile("(?<=[a-zA-Z])\'ll")
# to find the abbreviation of am
pat_am = re.compile("(?<=[I|i])\'m")
# to find the abbreviation of are
pat_are = re.compile("(?<=[a-zA-Z])\'re")
# to find the abbreviation of have
pat_ve = re.compile("(?<=[a-zA-Z])\'ve")

lmtzr = WordNetLemmatizer()

global index_word
global can_use_online_translate

def get_words(file):
    with open(file, encoding='utf-8') as f:
        words_box = []
        pat = re.compile(r'[^a-zA-Z \']+')
        for line in f:
            # if re.match(r'[a-zA-Z]*',line):
            #    words_box.extend(line.strip().strip('\'\"\.,').lower().split())
            # words_box.extend(pat.sub(' ', line).strip().lower().split())
            words_box.extend(merge(replace_abbreviations(line).split()))
    return collections.Counter(words_box)


def merge(words):
    new_words = []
    for word in words:
        if word:
            # tag is like [('bigger', 'JJR')]
            tag = nltk.pos_tag(word_tokenize(word))
            pos = get_wordnet_pos(tag[0][1])
            if pos:
                lemmatized_word = lmtzr.lemmatize(word, pos)
                new_words.append(lemmatized_word)
            else:
                new_words.append(word)
    return new_words


def get_wordnet_pos(treebank_tag):
    if treebank_tag.startswith('J'):
        return nltk.corpus.wordnet.ADJ
    elif treebank_tag.startswith('V'):
        return nltk.corpus.wordnet.VERB
    elif treebank_tag.startswith('N'):
        return nltk.corpus.wordnet.NOUN
    elif treebank_tag.startswith('R'):
        return nltk.corpus.wordnet.ADV
    else:
        return ''


def replace_abbreviations(text):
    new_text = text
    new_text = pat_letter.sub(' ', text).strip().lower()
    new_text = pat_is.sub(r"\1 is", new_text)
    new_text = pat_s.sub("", new_text)
    new_text = pat_s2.sub("", new_text)
    new_text = pat_not.sub(" not", new_text)
    new_text = pat_would.sub(" would", new_text)
    new_text = pat_will.sub(" will", new_text)
    new_text = pat_am.sub(" am", new_text)
    new_text = pat_are.sub(" are", new_text)
    new_text = pat_ve.sub(" have", new_text)
    new_text = new_text.replace('\'', ' ')
    return new_text


# def append_ext(words):
#     new_words = []
#     for item in words:
#         word, count = item
#         # tag is like [('bigger', 'JJR')]
#         tag = nltk.pos_tag(word_tokenize(word))[0][1]
#         new_words.append((word, count, tag))
#     return new_words


# def write_to_file(words, file='results.txt'):
#     f = open(file, 'w')
#     for item in words:
#         for field in item:
#             f.write(str(field)+',')
#         f.write('\n')

#################Translate################################
def fetch(query_str):
    query = {'q': "".join(query_str)}  # list --> str: "".join(list)
    url = 'https://fanyi.youdao.com/openapi.do?keyfrom=11pegasus11&key=273646050&type=data&doctype=json&version=1.1&' + \
          urlencode(query)

    proxies = {'http': 'http://127.0.0.1:1080'}

    response = urlopen(url, timeout=3)
    html = response.read().decode('utf-8')
    return html


def parse(html, word, count):
    return_ret = ""
    d = json.loads(html)
    global index_word
    try:
        if d.get('errorCode') == 0:
            explains = d.get('basic').get('explains')
            result = str(explains).replace('\'', "").replace(
                '[', "").replace(']', "")  # .replace真好用~
            sheet.cell(row=index_word, column=1).value = word
            sheet.cell(row=index_word, column=2).value = count
            sheet.cell(row=index_word, column=3).value = result
            if word in eudic_dict:
                sheet.cell(row=index_word, column=1).font = color_font
                # sheet.cell(row=index_word, column=2).font = color_font
                # sheet.cell(row=index_word, column=3).font = color_font

            index_word = index_word + 1

            for i in explains:
                print(i)

            return_ret = result
        else:
            print('无法翻译!****')
            if show_translate_error_word:
                sheet.cell(row=index_word, column=1).value = word
                if word in eudic_dict:
                    sheet.cell(row=index_word, column=1).font = color_font
                # sheet.cell(row=index_word, column=3).value = ' '  # 若无法翻译，则空出来
                index_word = index_word + 1
    except:
        print('****翻译出错!')  # 若无法翻译，则空出来
        if show_translate_error_word:
            sheet.cell(row=index_word, column=1).value = word
            if word in eudic_dict:
                sheet.cell(row=index_word, column=1).font = color_font
            # sheet.cell(row=index_word, column=3).value = ' '
            index_word = index_word + 1

    return return_ret




def local_translate(word, count):
    global index_word

    sheet.cell(row=index_word, column=1).value = word
    if word in eudic_dict:
        sheet.cell(row=index_word, column=1).font = color_font

    sheet.cell(row=index_word, column=2).value = count

    global can_use_online_translate
    if word in local_dict.keys():
        sheet.cell(row=index_word, column=3).value = local_dict[word]
        index_word = index_word + 1

    elif can_use_online_translate:
        try:
            online_result = parse(fetch(word), word, count)
            # add_local_dict_ret=""
            if online_result != "":
                add_local_dict_ret = word + "####1::" + online_result + '\n'

            else:  # 一次翻译不出，第二次也翻译不出，直接放入词典，避免在线查找
                add_local_dict_ret = word + "####-1::" + " " + '\n'  # -1表示即使在线查找也没有结果
            local_dict_file.write(add_local_dict_ret)
        except:
            print("online parse word error: " + word)
            print("youdao translate online limit, can not use online ")
            can_use_online_translate = False
            # out.save(outputFile)


def translate_and_write_to_file(words):
    num = 1
    for item in words:
        word, count = item
        if len(word) <= num_char_fliter or count <= rate_fliter:
            continue
        if word in filter_dict:
            continue
        if (word != None):
            print('正在翻译第', end='')
            print(num, end='')
            print('个单词')
            print(word)
            if use_online_translate:
                try:
                    parse(fetch(word), word, count)
                except:
                    print("parse word error: " + word)
                    out.save(outputFile)
            else:
                local_translate(word, count)

            num += 1
            print()
        else:
            print('翻译结束！')
            break


def get_filter_dict(filterNames):
    for filterName in filterNames:
        filterName = "data/" + filterName
        file = open(filterName, "r", encoding="utf-8")
        txt_string = file.read()
        txt_string_list = txt_string.split("\n")
        for temp in txt_string_list:
            filter_dict.add(temp)
        file.close()


def get_eudic_dict(eudicWordFile):
    eudicWordFile = "data/" + eudicWordFile
    file = open(eudicWordFile, "r", encoding="utf-8")
    txt_string = file.read()
    txt_string_list = txt_string.split("\n")

    for temp in txt_string_list:
        eudic_dict.add(temp)

    file.close()


def get_local_dict():
    file = open(localDictFile, 'r', encoding='utf-8')
    local_dict_list = file.read().split('\n')
    # print(local_dict_list[36325])
    # print(local_dict_list[36324])
    local_dict_list = local_dict_list[:-1]
    # print(local_dict_list[-1])
    for line in local_dict_list:
        temp = line.split("####")
        temp2 = temp[1].split("::")
        if temp2[0] != '0':
            local_dict[temp[0]] = temp2[1]

    file.close()


if __name__ == '__main__':
    # inpuFiles = ["1.txt", "17.txt", "大空头字幕.srt", "thinking in java.txt"]
    inpuFiles = ["thinking in java.txt"]
    filterFiles = ["fliter_city.txt", "fliter_first_name.txt", "fliter_second_name.txt", "fliter_simple_words.txt",
                   "fliter_others.txt"]
    eudicWordFile = "eudic_words.txt"
    num_char_fliter = 3  # 长度小于等于num_char_fliter不会显示
    rate_fliter = 0  # 出现次数小于rate_fliter，不显示
    use_online_translate = False  # if fase, first use local dict, then use online dict
    show_translate_error_word = True

    localDictFile = "data/local_dict.txt"

    ##########################
    global index_word
    global can_use_online_translate
    for inputFile in inpuFiles:
        index_word = 1
        can_use_online_translate = True
        print(inputFile + " start...")
        outputFile = inputFile + "_result.xlsx"
        inputFile = "data/" + inputFile

        local_dict = dict()
        if not use_online_translate:
            get_local_dict()

        local_dict_file = open(localDictFile, 'a', encoding='utf-8')  # if online trans have, add to local dict

        if os.path.isfile(outputFile):
            os.remove(outputFile)
        print("counting...")
        words = get_words(inputFile)
        print("there are " + str(len(words)) + " words")
        print("translate and writing file...")
        out = Workbook()
        sheet = out.active

        filter_dict = set()
        eudic_dict = set()
        get_filter_dict(filterFiles)
        get_eudic_dict(eudicWordFile)
        color_font = Font(color=RED)

        translate_and_write_to_file(words.most_common())

        local_dict_file.close()
        out.save(outputFile)
