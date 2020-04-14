# -*- coding: UTF-8 -*-
import openpyxl
import os.path


def writeCSVFile(input_file, sheet_name):
    output_file = input_file+".csv"
    input_file = input_file+".xlsx"
    if os.path.isfile(output_file):
        os.remove(output_file)

    workbook = openpyxl.load_workbook(input_file)
    worksheet = workbook[sheet_name]
    ret = ""
    for i in range(2, worksheet.max_row):
        ret = ret+str(worksheet.cell(i, 1).value)+"," + \
            str(worksheet.cell(i, 3).value)+"\n"

    file = open(output_file, 'w')
    file.write(ret)
    file.close()


def main():
    ###########################
    input_file = "think in java word lists"
    sheet_name = "筛选单词"
    ###########################
    print("开始写入CSV文件")
    writeCSVFile(input_file, sheet_name)
    print("完成")


if __name__ == "__main__":
    main()
